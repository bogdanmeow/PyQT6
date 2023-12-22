import tkinter as tk
from openpyxl import load_workbook


def check_employee():
    employee_name = name_entry.get()

    # Загрузка таблицы Excel
    workbook = load_workbook('employees.xlsx')
    sheet = workbook.active

    # Поиск сотрудника в таблице
    for row in sheet.iter_rows(min_row=2, values_only=True):
        last_name, first_name, middle_name = row[0], row[1], row[2]
        if first_name == employee_name:
            result_label.config(text="Проход разрешен")
            return

    # Если сотрудник не найден
    result_label.config(text="Проход запрещен")


# Создание окна
window = tk.Tk()
window.title("Проверка сотрудника")

# Поле для ввода имени сотрудника
name_label = tk.Label(window, text="Введите имя сотрудника:")
name_label.pack()

name_entry = tk.Entry(window)
name_entry.pack()

# Кнопка "Проверить"
check_button = tk.Button(window, text="Проверить", command=check_employee)
check_button.pack()

# Метка для вывода результата
result_label = tk.Label(window, text="")
result_label.pack()

window.mainloop()